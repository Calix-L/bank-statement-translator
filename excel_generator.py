"""Generate formatted Excel from structured data."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from statement_structurer import PDF_HEADER_MAP_EN, PDF_HEADER_ORDER, PDF_HEADER_ORDER_EN


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class ExcelGenerator:
    @staticmethod
    def generate(translated_df: pd.DataFrame, raw_df: pd.DataFrame | None = None) -> bytes:
        translated_df = ExcelGenerator._align_pdf_columns(translated_df)
        translated_df = translated_df.rename(columns=PDF_HEADER_MAP_EN)[PDF_HEADER_ORDER_EN]
        
        # Ensure raw_df has correct column names
        if raw_df is not None:
            raw_df = ExcelGenerator._align_pdf_columns(raw_df)
            # Use PDF_HEADER_ORDER as column names (Chinese)
            raw_df.columns = PDF_HEADER_ORDER

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            translated_df.to_excel(writer, index=False, sheet_name="Translated")
            if raw_df is not None and not raw_df.empty:
                raw_df.to_excel(writer, index=False, sheet_name="Raw")

            for ws in writer.book.worksheets:
                ExcelGenerator._style_sheet(ws)

        return output.getvalue()

    @staticmethod
    def generate_batch(results: list[dict], output_filename: str = "batch_output.xlsx") -> bytes:
        """Generate Excel file with multiple sheets from batch processing results.
        
        Args:
            results: List of dicts with keys 'name', 'translated_df', 'raw_df'
            
        Returns:
            Excel file bytes
        """
        output = BytesIO()
        has_data = False

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for result in results:
                name = result.get("name", "Unknown")
                translated_df = result.get("translated_df")
                raw_df = result.get("raw_df")

                if translated_df is None or translated_df.empty:
                    continue

                has_data = True

                trans_sheet = ExcelGenerator._align_pdf_columns(translated_df)
                trans_sheet = trans_sheet.rename(columns=PDF_HEADER_MAP_EN)[PDF_HEADER_ORDER_EN]
                trans_sheet.to_excel(writer, index=False, sheet_name=f"{name}_Translated")

                if raw_df is not None and not raw_df.empty:
                    raw_sheet = ExcelGenerator._align_pdf_columns(raw_df)
                    raw_sheet.columns = PDF_HEADER_ORDER
                    raw_sheet.to_excel(writer, index=False, sheet_name=f"{name}_Raw"[:31])

            if not has_data:
                pd.DataFrame({"Status": ["No batch results"]}).to_excel(
                    writer,
                    index=False,
                    sheet_name="Summary",
                )

            for ws in writer.book.worksheets:
                ExcelGenerator._style_sheet(ws)

        return output.getvalue()

    @staticmethod
    def _align_pdf_columns(df: pd.DataFrame) -> pd.DataFrame:
        aligned = df.copy()
        for col in PDF_HEADER_ORDER:
            if col not in aligned.columns:
                aligned[col] = ""
        return aligned[PDF_HEADER_ORDER]

    @staticmethod
    def _style_sheet(ws) -> None:
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)
